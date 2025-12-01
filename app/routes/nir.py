from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, send_file
from flask_login import login_required, current_user
from app.models import db, Nir, User, NirProcedure, NirSectionStatus
from app.procedures_models import Procedure
from app.utils.rbac_permissions import require_permission, require_sector
from app.routes.util import format_date_filter
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
import os
import unicodedata
from werkzeug.utils import secure_filename

nir_bp = Blueprint('nir', __name__, template_folder='../templates')

#<!--- Funções Auxiliares --->

def parse_datetime_local(value):
    """Parse datetime-local input (YYYY-MM-DDTHH:MM) or date input (YYYY-MM-DD)"""
    if not value:
        return None
    try:
        # Try datetime-local format first (YYYY-MM-DDTHH:MM)
        return datetime.strptime(value, '%Y-%m-%dT%H:%M')
    except ValueError:
        try:
            # Fallback to date format (YYYY-MM-DD) at midnight
            return datetime.strptime(value, '%Y-%m-%d')
        except ValueError:
            return None

def create_iter_pages_function(pages, current_page):
    def _iter_pages(left_edge=1, right_edge=1, left_current=2, right_current=2):
        last = 0
        for num in range(1, pages + 1):
            if num <= left_edge or \
               (num > current_page - left_current - 1 and num < current_page + right_current) or \
               num > pages - right_edge:
                if last + 1 != num:
                    yield None
                yield num
                last = num
    return _iter_pages

def get_nir_phase(record):
    config = record.get_section_control_config()
    nir_sections = [s for s, sec in config.items() if sec == 'NIR']
    
    initial_sections = ['dados_paciente', 'dados_internacao_iniciais', 'agendamento_inicial']
    alta_sections = ['dados_alta_finais']
    
    initial_sections = [s for s in initial_sections if s in nir_sections]
    alta_sections = [s for s in alta_sections if s in nir_sections]

    status_map = {s.section_name: s.status for s in record.section_statuses if s.responsible_sector == 'NIR'}
    def sections_complete(lst):
        if not lst:
            return True
        return all(status_map.get(sec) == 'PREENCHIDO' for sec in lst)

    effective_entry = record.get_effective_entry_type()
    progress = record.get_sector_progress()
    surgery_progress = progress.get('CENTRO_CIRURGICO')
    surgery_complete = (surgery_progress and surgery_progress.get('status') == 'CONCLUIDO') if surgery_progress else True
    billing_progress = progress.get('FATURAMENTO', {})
    billing_complete = billing_progress.get('status') == 'CONCLUIDO'

    if record.admission_type == 'CLINICO':
        
        initial_complete = sections_complete(initial_sections)
        final_complete = sections_complete(alta_sections)
        
        for sec in initial_sections:
            sec_status = status_map.get(sec, 'N/A')
        
        for sec in alta_sections:
            sec_status = status_map.get(sec, 'N/A')
        
        if not initial_complete:
            return dict(phase='INITIAL', locked=False, waiting_for=None, show_sections=initial_sections)
        
        if initial_complete and not final_complete:
            return dict(phase='FINAL', locked=False, waiting_for=None, show_sections=alta_sections)
        
        if initial_complete and final_complete and not billing_complete:
            return dict(phase='LOCKED_AFTER', locked=True, waiting_for='FATURAMENTO', show_sections=[])
        
        return dict(phase='LOCKED_AFTER', locked=True, waiting_for=None, show_sections=[])

    if effective_entry in ('URGENCIA', 'ELETIVO') or record.admission_type == 'CIRURGICO':
        initial_complete = sections_complete(initial_sections)
        final_complete = sections_complete(alta_sections)
        if not initial_complete:
            return dict(phase='INITIAL', locked=False, waiting_for=None, show_sections=initial_sections)
        if initial_complete and not surgery_complete:
            return dict(phase='LOCKED_WAIT_SURGERY', locked=True, waiting_for='CENTRO CIRÚRGICO', show_sections=[])
        if initial_complete and surgery_complete and not final_complete:
            return dict(phase='FINAL', locked=False, waiting_for=None, show_sections=alta_sections)
        if initial_complete and surgery_complete and final_complete and not billing_complete:
            return dict(phase='LOCKED_AFTER', locked=True, waiting_for='FATURAMENTO', show_sections=[])
        return dict(phase='LOCKED_AFTER', locked=True, waiting_for=None, show_sections=[])
    else:
        initial_complete = sections_complete(initial_sections)
        
        if not initial_complete:
            return dict(phase='INITIAL', locked=False, waiting_for=None, show_sections=initial_sections)
        
        all_complete = sections_complete(nir_sections)
        if not all_complete:
            return dict(phase='FULL', locked=False, waiting_for=None, show_sections=nir_sections)
        if all_complete and not billing_complete:
            return dict(phase='LOCKED_AFTER_FULL', locked=True, waiting_for='FATURAMENTO', show_sections=[])
        return dict(phase='LOCKED_AFTER_FULL', locked=True, waiting_for=None, show_sections=[])

def initialize_section_statuses(nir):
    config = nir.get_section_control_config()
    
    for section_name, responsible_sector in config.items():
        existing = NirSectionStatus.query.filter_by(
            nir_id=nir.id,
            section_name=section_name
        ).first()
        
        if not existing:
            section_status = NirSectionStatus(
                nir_id=nir.id,
                section_name=section_name,
                responsible_sector=responsible_sector,
                status='PENDENTE'
            )
            db.session.add(section_status)
        else:
            if existing.responsible_sector != responsible_sector:
                existing.responsible_sector = responsible_sector

def get_user_sector(user):
    user_sectors = [role.sector for role in user.roles if role.sector]
    
    if 'CENTRO_CIRURGICO' in user_sectors:
        return 'CENTRO_CIRURGICO'
    elif 'FATURAMENTO' in user_sectors:
        return 'FATURAMENTO'
    elif any(sector in ['ENFERMAGEM', 'MEDICINA', 'RECEPCAO'] for sector in user_sectors):
        return 'NIR'
    else:
        return 'NIR'

def update_section_status(nir_id, section_name, user_id):
    section_status = NirSectionStatus.query.filter_by(
        nir_id=nir_id,
        section_name=section_name
    ).first()
        
    if section_status:
        old_status = section_status.status
        section_status.status = 'PREENCHIDO'
        section_status.filled_by_user_id = user_id
        section_status.filled_at = datetime.utcnow()
        db.session.commit()

def _compute_global_info(rec):
    if rec.status == 'EM_OBSERVACAO':
        hours = rec.observation_hours_elapsed()
        return 'EM_OBSERVACAO', f'Em observação há {hours:.1f}h'
    elif rec.status == 'AGUARDANDO_DECISAO':
        hours = rec.observation_hours_elapsed()
        return 'AGUARDANDO_DECISAO', f'Aguardando decisão ({hours:.1f}h)'
    
    try:
        if (rec.cancelled or '').upper() == 'SIM' or rec.status == 'CANCELADO':
            return 'CANCELADO', 'Fluxo cancelado'
    except Exception:
        pass
    prog = rec.get_sector_progress()
    billing = (prog.get('FATURAMENTO') or {}).get('status')
    if billing == 'CONCLUIDO':
        return 'CONCLUIDO', 'Fluxo concluído'
    current_sector = rec.get_next_available_sector()
    sector_labels = {
        'NIR': 'NIR',
        'CENTRO_CIRURGICO': 'Centro Cirúrgico',
        'FATURAMENTO': 'Faturamento',
    }
    if current_sector == 'FATURAMENTO':
        return 'EM_ANDAMENTO', f"Aguardando {sector_labels.get(current_sector, current_sector or '')}"
    if current_sector == 'NIR':
        effective_entry = rec.get_effective_entry_type()
        config = rec.get_section_control_config()
        nir_sections = [s for s, sec in config.items() if sec == 'NIR']
        alta_sections = [s for s in nir_sections if 'alta' in s]
        initial_sections = [s for s in nir_sections if s not in alta_sections]
        status_map = {s.section_name: s.status for s in rec.section_statuses if s.responsible_sector == 'NIR'}
        def sections_complete(lst):
            if not lst:
                return True
            return all(status_map.get(sec) == 'PREENCHIDO' for sec in lst)
        status = (prog.get('NIR') or {}).get('status') or 'PENDENTE'
        if effective_entry in ('URGENCIA', 'ELETIVO') and sections_complete(initial_sections) and not sections_complete(alta_sections):
            return status, 'Aguardando dados de Alta'
        return status, 'Aguardando NIR'
    if current_sector:
        status = (prog.get(current_sector) or {}).get('status') or 'PENDENTE'
        hint = f"Aguardando {sector_labels.get(current_sector, current_sector)}"
        return status, hint
    return 'CONCLUIDO', 'Fluxo concluído'

#<!--- Lista de Registros NIR --->
@nir_bp.route("/nir")
@login_required
def list_records():
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 10, type=int)
    if per_page <= 0:
        per_page = 10
    if per_page > 100:
        per_page = 100
    is_ajax = request.args.get('ajax') == '1' or request.headers.get('X-Requested-With') == 'XMLHttpRequest'

    search = request.args.get('search', '').strip()
    entry_type = request.args.get('entry_type', '').strip()
    admission_type = request.args.get('admission_type', '').strip()
    discharge_type = request.args.get('discharge_type', '').strip()
    discharge_type = request.args.get('discharge_type', '').strip()
    is_palliative = request.args.get('is_palliative', '').strip()
    origin = request.args.get('origin', '').strip()
    recurso = request.args.get('recurso', '').strip()
    responsible_doctor = request.args.get('responsible_doctor', '').strip()
    start_date = request.args.get('start_date', '').strip()
    end_date = request.args.get('end_date', '').strip()
    sector = request.args.get('sector', '').strip()
    sector_progress = request.args.get('sector_progress', '').strip()

    query = Nir.query.order_by(Nir.creation_date.desc())

    if search:
        query = query.filter(
            (Nir.patient_name.ilike(f'%{search}%')) |
            (Nir.susfacil_protocol.ilike(f'%{search}%'))
        )

    if entry_type:
        query = query.filter(Nir.entry_type.ilike(entry_type))

    if admission_type:
        query = query.filter(Nir.admission_type.ilike(admission_type))
    
    if discharge_type:
        query = query.filter(Nir.discharge_type.ilike(discharge_type))
    
    if is_palliative:
        if is_palliative == '1':
            query = query.filter(Nir.is_palliative == True)
        elif is_palliative == '0':
            query = query.filter(Nir.is_palliative == False)
    
    if origin:
        query = query.filter(Nir.admitted_from_origin == origin)
    
    if recurso:
        query = query.filter(Nir.recurso == recurso)

    if responsible_doctor:
        query = query.filter(Nir.responsible_doctor.ilike(f'%{responsible_doctor}%'))

    if start_date:
        try:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
            query = query.filter(Nir.admission_date >= start_date_obj)
        except ValueError:
            flash('Data inicial inválida', 'warning')

    if end_date:
        try:
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
            query = query.filter(Nir.admission_date <= end_date_obj)
        except ValueError:
            flash('Data final inválida', 'warning')

    all_basic_records = query.all()

    for rec in all_basic_records:
        try:
            st, hint = _compute_global_info(rec)
            setattr(rec, '_global_status', st)
            setattr(rec, '_global_status_hint', hint)
        except Exception:
            setattr(rec, '_global_status', None)
            setattr(rec, '_global_status_hint', None)

    global_for_counts = all_basic_records

    if sector or sector_progress:
        def matches_sector(rec):
            prog = rec.get_sector_progress()
            sect = sector if sector else None
            if sect and sect not in prog:
                return False
            if sect and sector_progress:
                return (prog.get(sect) or {}).get('status') == sector_progress
            if sect and not sector_progress:
                return (prog.get(sect) or {}).get('status') in ('PENDENTE', 'EM_ANDAMENTO')
            if not sect and sector_progress:
                return (getattr(rec, '_global_status', None) or 'PENDENTE') == sector_progress
            return True
        display_records = [r for r in all_basic_records if matches_sector(r)]
    else:
        display_records = all_basic_records

    def _status_sort_key(rec):
        st = getattr(rec, '_global_status', None) or 'PENDENTE'
        if st == 'PENDENTE':
            pri = 0
        elif st == 'EM_ANDAMENTO':
            pri = 1
        elif st == 'CONCLUIDO':
            pri = 2
        else:
            pri = 3
        created_ts = 0
        try:
            if rec.creation_date:
                created_ts = rec.creation_date.timestamp()
        except Exception:
            pass
        return (pri, -created_ts)

    try:
        display_records.sort(key=_status_sort_key)
    except Exception:
        pass

    from math import ceil
    total_display = len(display_records)
    start_idx = (page - 1) * per_page
    end_idx = start_idx + per_page
    page_items = display_records[start_idx:end_idx]

    from types import SimpleNamespace
    pages = ceil(total_display / per_page) if total_display else 1
    records = SimpleNamespace(
        items=page_items,
        total=total_display,
        pages=pages,
        page=page,
        has_prev=page > 1,
        has_next=page < pages,
        prev_num=page - 1 if page > 1 else None,
        next_num=page + 1 if page < pages else None,
        iter_pages=create_iter_pages_function(pages, page)
    )

    filtered_records = global_for_counts
    
    stats_counts = {
        'total': len(filtered_records),
        'pendentes': 0,
        'andamento': 0,
        'concluidos': 0,
        'cancelados': 0,
    }
    for r in filtered_records:
        status = getattr(r, '_global_status', None) or 'PENDENTE'
        if status == 'PENDENTE':
            stats_counts['pendentes'] += 1
        elif status == 'EM_ANDAMENTO':
            stats_counts['andamento'] += 1
        elif status == 'CONCLUIDO':
            stats_counts['concluidos'] += 1
        elif status == 'CANCELADO':
            stats_counts['cancelados'] += 1

    entry_types = list(set(r.entry_type for r in all_basic_records if r.entry_type))
    admission_types = list(set(r.admission_type for r in all_basic_records if r.admission_type))
    discharge_types = list(set(r.discharge_type for r in all_basic_records if r.discharge_type))
    
    entry_types.sort()
    admission_types.sort()
    discharge_types.sort()

    filters = {
        'search': search,
        'entry_type': entry_type,
        'admission_type': admission_type,
        'discharge_type': discharge_type,
        'is_palliative': is_palliative,
        'origin': origin,
        'recurso': recurso,
        'responsible_doctor': responsible_doctor,
        'start_date': start_date,
        'end_date': end_date,
        'sector': sector,
        'sector_progress': sector_progress,
        'per_page': per_page
    }

    if is_ajax:
        return render_template('nir/_records_table.html', 
                             records=records, 
                             entry_types=entry_types,
                             admission_types=admission_types,
                             discharge_types=discharge_types,
                             filters=filters,
                             stats_counts=stats_counts)

    return render_template('nir/list_records.html', 
                         records=records, 
                         entry_types=entry_types,
                         admission_types=admission_types,
                         discharge_types=discharge_types,
                         filters=filters,
                         stats_counts=stats_counts)

#<!--- ROTAS PARA FLUXO DE OBSERVAÇÃO --->

#<!--- Evoluir Observação para Internação --->
@nir_bp.route("/nir/observacao/<int:record_id>/evoluir", methods=['POST'])
@login_required
@require_sector('NIR')
@require_permission('editar-registro-nir')
def evolve_observation(record_id):
    """Evolui uma observação para internação normal"""
    record = Nir.query.get_or_404(record_id)
    
    if record.status not in ('EM_OBSERVACAO', 'AGUARDANDO_DECISAO'):
        flash('Este registro não está em observação.', 'warning')
        return redirect(url_for('nir.record_details', record_id=record_id))
    
    try:
        admission_date_str = request.form.get('admission_date')
        if admission_date_str:
            record.admission_date = parse_datetime_local(admission_date_str)
            if record.admission_date:
                record.day = record.admission_date.day
                record.month = record.admission_date.strftime('%B')
        elif not record.admission_date and record.fa_datetime:
            record.admission_date = record.fa_datetime
            record.day = record.admission_date.day
            record.month = record.admission_date.strftime('%B')
        
        record.entry_type = request.form.get('entry_type')
        record.admission_type = request.form.get('admission_type')
        record.admitted_from_origin = request.form.get('admitted_from_origin')
        
        scheduling_date_str = request.form.get('scheduling_date')
        if scheduling_date_str:
            from datetime import datetime
            record.scheduling_date = datetime.strptime(scheduling_date_str, '%Y-%m-%d').date()
        
        if record.entry_type == 'ELETIVO':
            susfacil_accepted = request.form.get('susfacil_accepted') == 'on'
            record.susfacil_accepted = susfacil_accepted
            if susfacil_accepted:
                susfacil_datetime_str = request.form.get('susfacil_accept_datetime')
                if susfacil_datetime_str:
                    from datetime import datetime
                    record.susfacil_accept_datetime = datetime.strptime(susfacil_datetime_str, '%Y-%m-%dT%H:%M')
        
        if record.admission_type == 'CLINICO':
            record.is_palliative = request.form.get('is_palliative') == 'on'
            record.responsible_doctor = request.form.get('responsible_doctor')
            record.surgical_specialty = request.form.get('surgical_specialty')
            record.main_cid = request.form.get('main_cid')
            
            procedure_codes = request.form.getlist('procedure_codes[]')
            procedure_descriptions = request.form.getlist('procedure_descriptions[]')
            
            if procedure_codes and procedure_descriptions:
                NirProcedure.query.filter_by(nir_id=record.id).delete()
                
                for idx, (code, description) in enumerate(zip(procedure_codes, procedure_descriptions)):
                    procedure = NirProcedure(
                        nir_id=record.id,
                        code=code,
                        description=description,
                        sequence=idx + 1,
                        is_primary=(idx == 0)
                    )
                    db.session.add(procedure)
        
        record.evolve_to_admission()
        
        initialize_section_statuses(record)
        
        config = record.get_section_control_config()
        nir_sections = [section_name for section_name, sector in config.items() if sector == 'NIR']
        
        alta_sections = [s for s in nir_sections if 'alta' in s.lower()]
        initial_sections = [s for s in nir_sections if s not in alta_sections]
        
        for section_name in initial_sections:
            update_section_status(record.id, section_name, current_user.id)
        
        db.session.commit()
        
        flash(f'Paciente {record.patient_name} evoluído para internação com sucesso! Dados iniciais preenchidos.', 'success')
        return redirect(url_for('nir.sector_nir_list'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao evoluir observação: {str(e)}', 'danger')
        return redirect(url_for('nir.sector_nir_list'))

#<!--- Cancelar Observação --->
@nir_bp.route("/nir/observacao/<int:record_id>/cancelar", methods=['POST'])
@login_required
@require_sector('NIR')
@require_permission('editar-registro-nir')
def cancel_observation(record_id):
    """Cancela uma observação (paciente recebeu alta antes de 24h)"""
    record = Nir.query.get_or_404(record_id)
    
    if record.status not in ('EM_OBSERVACAO', 'AGUARDANDO_DECISAO'):
        flash('Este registro não está em observação.', 'warning')
        return redirect(url_for('nir.record_details', record_id=record_id))
    
    try:
        reason = request.form.get('cancellation_reason', 'Alta antes da internação')
        
        record.cancel_observation(reason)
        db.session.commit()
        
        flash(f'Observação #{record.id} cancelada com sucesso.', 'info')
        return redirect(url_for('nir.sector_nir_list'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao cancelar observação: {str(e)}', 'danger')
        return redirect(url_for('nir.sector_nir_list'))


#<!--- Rota de Busca de Procedimentos --->
@nir_bp.route("/nir/search_procedures", methods=['GET'])
@login_required
def search_procedures():
    """Busca procedimentos por código ou descrição"""
    try:
        from app.procedures_models import Procedure, ProcedureCid, Cid
        
        query = request.args.get('q', '').strip()
        
        if not query or len(query) < 2:
            return jsonify({'results': []})
        
        procedures = Procedure.query.filter(
            db.or_(
                Procedure.code.ilike(f'%{query}%'),
                Procedure.description.ilike(f'%{query}%')
            )
        ).filter_by(is_active=True).limit(10).all()
        
        results = []
        for proc in procedures:
            procedure_cids = ProcedureCid.query.filter_by(procedure_code=proc.code).all()
            
            cid_list = []
            for proc_cid in procedure_cids:
                if proc_cid.cid and proc_cid.cid.is_active:
                    cid_list.append({
                        'code': proc_cid.cid.code,
                        'description': proc_cid.cid.description
                    })
            
            results.append({
                'code': proc.code,
                'description': proc.description,
                'cids': cid_list
            })
        
        return jsonify({'results': results})
        
    except Exception as e:
        print(f"Erro ao buscar procedimentos: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e), 'results': []}), 500


#<!--- Endpoint para Verificar Pendências (Sistema de Notificações) --->
@nir_bp.route("/nir/check-pending-notifications", methods=['GET'])
@login_required
def check_pending_notifications():
    """Verifica solicitações que precisam de atenção"""
    try:
        from datetime import datetime, timedelta
        
        notifications = []
        
        time_24h_ago = datetime.now() - timedelta(hours=24)
        
        pending_decision = Nir.query.filter(
            Nir.status == 'AGUARDANDO_DECISAO',
            Nir.fa_datetime <= time_24h_ago
        ).count()
        
        if pending_decision > 0:
            notifications.append({
                'type': 'warning',
                'priority': 'high',
                'title': 'Solicitações Aguardando Decisão',
                'message': f'{pending_decision} {"solicitação" if pending_decision == 1 else "solicitações"} aguardando decisão há mais de 24 horas.',
                'action_url': url_for('nir.sector_nir_list'),
                'action_text': 'Ver Solicitações',
                'icon': 'bi-clock-history'
            })
        
        time_22h_ago = datetime.now() - timedelta(hours=22)
        
        observation_critical = Nir.query.filter(
            Nir.status == 'EM_OBSERVACAO',
            Nir.fa_datetime <= time_22h_ago,
            Nir.fa_datetime > time_24h_ago
        ).count()
        
        if observation_critical > 0:
            notifications.append({
                'type': 'info',
                'priority': 'medium',
                'title': 'Observações Próximas do Limite',
                'message': f'{observation_critical} {"paciente" if observation_critical == 1 else "pacientes"} em observação próximo(s) das 24 horas.',
                'action_url': url_for('nir.sector_nir_list'),
                'action_text': 'Ver Observações',
                'icon': 'bi-hourglass-split'
            })
        
        return jsonify({
            'has_notifications': len(notifications) > 0,
            'count': len(notifications),
            'notifications': notifications
        })
        
    except Exception as e:
        print(f"Erro ao verificar pendências: {str(e)}")
        return jsonify({
            'has_notifications': False,
            'count': 0,
            'notifications': [],
            'error': str(e)
        }), 500


#<!--- Rota de Formulário e Criação de Registro NIR pelo Setor NIR --->
@nir_bp.route("/nir/setor/novo", methods=['GET', 'POST'])
@login_required
@require_sector('NIR')
@require_permission('criar-registro-nir')
def sector_new_record():
    user_sector = get_user_sector(current_user)
    if user_sector != 'NIR':
        flash('Acesso negado: apenas funcionários do NIR podem criar novos registros aqui', 'danger')
        return redirect(url_for('nir.my_work'))

    if request.method == 'GET':
        return render_template('nir/sector_new_record.html', user_sector=user_sector)

    try:
        record_type = request.form.get('record_type', 'admission')
        
        birth_date_str = request.form.get('birth_date')
        birth_date = datetime.strptime(birth_date_str, '%Y-%m-%d').date() if birth_date_str else None

        if record_type == 'observation':
            fa_datetime_str = request.form.get('fa_datetime')
            if not fa_datetime_str:
                flash('Por favor, informe o Horário FA (entrada na Fila de Atendimento).', 'warning')
                return redirect(url_for('nir.sector_new_record'))
            
            try:
                fa_datetime = datetime.strptime(fa_datetime_str, '%Y-%m-%dT%H:%M')
            except ValueError:
                flash('Horário FA inválido. Use o formato correto de data e hora.', 'danger')
                return redirect(url_for('nir.sector_new_record'))
            
            new_nir = Nir(
                patient_name=request.form.get('patient_name'),
                birth_date=birth_date,
                gender=request.form.get('gender'),
                susfacil=request.form.get('susfacil'),
                sus_number=request.form.get('sus_number'),
                is_palliative=request.form.get('is_palliative') == 'on',
                
                admission_date=None,
                entry_type=None,
                admission_type=None,
                admitted_from_origin=None,
                scheduling_date=None,
                
                status='EM_OBSERVACAO',
                observation_start_time=datetime.utcnow(),
                fa_datetime=fa_datetime,
                
                observation=request.form.get('observation'),
                operator_id=current_user.id
            )
            
            db.session.add(new_nir)
            db.session.commit()
            
            flash(f'Solicitação de observação #{new_nir.id} criada com sucesso.', 'success')
            return redirect(url_for('nir.sector_nir_list'))
        
        admission_date_str = request.form.get('admission_date')
        admission_date = parse_datetime_local(admission_date_str)

        scheduling_date_str = request.form.get('scheduling_date')
        scheduling_date = datetime.strptime(scheduling_date_str, '%Y-%m-%d').date() if scheduling_date_str else None

        discharge_date_str = request.form.get('discharge_date')
        discharge_date = parse_datetime_local(discharge_date_str)

        total_days = None
        if admission_date and discharge_date:
            total_days = (discharge_date.date() - admission_date.date()).days if hasattr(admission_date, 'date') else (discharge_date - admission_date).days

        entry_type_form = request.form.get('entry_type')
        admission_type_form = request.form.get('admission_type')
        aih_form_value = request.form.get('aih')
        if admission_type_form == 'CIRURGICO':
            if 'aih' in request.form and (not aih_form_value or not aih_form_value.strip()):
                flash('AIH é obrigatório para internação cirúrgica.', 'danger')
                return redirect(url_for('nir.sector_new_record'))
        
        if aih_form_value and aih_form_value.strip():
            existing_aih = Nir.query.filter_by(aih=aih_form_value.strip()).first()
            if existing_aih:
                flash(f'Já existe uma solicitação com o AIH {aih_form_value}. O número de AIH deve ser único.', 'danger')
                return redirect(url_for('nir.sector_new_record'))

        susfacil_accepted = request.form.get('susfacil_accepted') == 'on'
        susfacil_accept_datetime_str = request.form.get('susfacil_accept_datetime')
        susfacil_accept_datetime = None
        susfacil_protocol = None
        
        if entry_type_form == 'ELETIVO':
            if not susfacil_accepted:
                flash('Para internações eletivas, você deve confirmar que realizou o aceite no SUSFACIL.', 'warning')
                return redirect(url_for('nir.sector_new_record'))
            
            if not susfacil_accept_datetime_str:
                flash('Por favor, informe a data e hora do aceite no SUSFACIL.', 'warning')
                return redirect(url_for('nir.sector_new_record'))
            
            try:
                susfacil_accept_datetime = datetime.strptime(susfacil_accept_datetime_str, '%Y-%m-%dT%H:%M')
                import random
                dt = susfacil_accept_datetime
                random_suffix = f"{random.randint(0, 9999):04d}"
                susfacil_protocol = f"NIR-{dt.strftime('%Y%m%d-%H%M')}-{random_suffix}"
            except ValueError:
                flash('Data e hora do aceite SUSFACIL inválida.', 'danger')
                return redirect(url_for('nir.sector_new_record'))

        surgical_specialty_form = request.form.get('surgical_specialty') or None
        surgical_type_form = request.form.get('surgical_type') or None
        anesthetist_form = request.form.get('anesthetist') or None
        anesthesia_form = request.form.get('anesthesia') or None
        auxiliary_form = None if request.form.get('auxiliary_nao_aplica') else (request.form.get('auxiliary') or None)
        pediatrics_form = None if request.form.get('pediatrics_nao_aplica') else (request.form.get('pediatrics') or None)

        new_nir = Nir(
            patient_name=request.form.get('patient_name'),
            birth_date=birth_date,
            gender=request.form.get('gender'),
            susfacil=request.form.get('susfacil'),
            sus_number=request.form.get('sus_number'),
            is_palliative=request.form.get('is_palliative') == 'on',
            admission_date=admission_date,
            entry_type=entry_type_form,
            admission_type=admission_type_form,
            admitted_from_origin=request.form.get('admitted_from_origin'),
            recurso=request.form.get('recurso'),
            procedure_code=None,
            surgical_description=None,
            responsible_doctor=request.form.get('responsible_doctor') if get_user_sector(current_user) == 'NIR' else None,
            main_cid=request.form.get('main_cid') if get_user_sector(current_user) == 'NIR' else None,
            aih=aih_form_value,
            susfacil_accepted=susfacil_accepted if entry_type_form == 'ELETIVO' else None,
            susfacil_accept_datetime=susfacil_accept_datetime if entry_type_form == 'ELETIVO' else None,
            susfacil_protocol=susfacil_protocol if entry_type_form == 'ELETIVO' else None,
            scheduling_date=scheduling_date if get_user_sector(current_user) == 'NIR' else None,
            discharge_type=request.form.get('discharge_type') if (entry_type_form == 'ELETIVO' and get_user_sector(current_user) == 'NIR') else None,
            discharge_date=discharge_date if (entry_type_form == 'ELETIVO' and get_user_sector(current_user) == 'NIR') else None,
            total_days_admitted=total_days if get_user_sector(current_user) == 'NIR' else None,
            cancelled=None,
            cancellation_reason=None,
            criticized=None,
            billed=None,
            status='PENDENTE',
            observation=request.form.get('observation'),
            surgical_specialty=surgical_specialty_form if entry_type_form == 'ELETIVO' and get_user_sector(current_user) == 'NIR' else None,
            auxiliary=auxiliary_form if entry_type_form == 'ELETIVO' and get_user_sector(current_user) == 'NIR' else None,
            anesthetist=anesthetist_form if entry_type_form == 'ELETIVO' and get_user_sector(current_user) == 'NIR' else None,
            anesthesia=anesthesia_form if entry_type_form == 'ELETIVO' and get_user_sector(current_user) == 'NIR' else None,
            pediatrics=pediatrics_form if entry_type_form == 'ELETIVO' and get_user_sector(current_user) == 'NIR' else None,
            surgical_type=surgical_type_form if entry_type_form == 'ELETIVO' and get_user_sector(current_user) == 'NIR' else None,
            day=admission_date.day if admission_date else None,
            month=admission_date.strftime('%B') if admission_date else None,
            operator_id=current_user.id
        )

        db.session.add(new_nir)
        db.session.flush()

        initialize_section_statuses(new_nir)

        codes = request.form.getlist('procedure_codes[]')
        is_primary_list = request.form.getlist('procedure_is_primary[]')
        
        seen = set()
        for idx, code in enumerate(codes):
            code_norm = (code or '').strip()
            if not code_norm or code_norm in seen:
                continue
            seen.add(code_norm)
            
            from app.procedures_models import Procedure
            procedure = Procedure.query.filter_by(code=code_norm).first()
            
            if procedure:
                is_primary = (is_primary_list[idx] == '1') if idx < len(is_primary_list) else (idx == 0)
                
                proc = NirProcedure(
                    nir_id=new_nir.id,
                    code=code_norm,
                    description=procedure.description,
                    sequence=idx,
                    is_primary=is_primary
                )
                db.session.add(proc)
                
                if is_primary:
                    new_nir.procedure_code = code_norm
                    new_nir.surgical_description = procedure.description

        if not seen:
            single_code = request.form.get('procedure_code')
            single_desc = request.form.get('surgical_description')
            if single_code and single_desc:
                new_nir.procedure_code = single_code
                new_nir.surgical_description = single_desc
                db.session.add(NirProcedure(
                    nir_id=new_nir.id,
                    code=single_code,
                    description=single_desc,
                    sequence=0,
                    is_primary=True
                ))

        if admission_type_form == 'CLINICO':
            responsible_doctor_clinico = request.form.get('responsible_doctor')
            main_cid_clinico = request.form.get('main_cid')
            surgical_specialty_clinico = request.form.get('surgical_specialty')
            
            if responsible_doctor_clinico:
                new_nir.responsible_doctor = responsible_doctor_clinico
            if main_cid_clinico:
                new_nir.main_cid = main_cid_clinico
            if surgical_specialty_clinico:
                new_nir.surgical_specialty = surgical_specialty_clinico

        db.session.commit()

        try:
            user_sector_created = get_user_sector(current_user)
            if user_sector_created == 'NIR':
                config_sections = new_nir.get_section_control_config()
                
                if new_nir.patient_name and new_nir.birth_date and new_nir.gender and new_nir.sus_number:
                    update_section_status(new_nir.id, 'dados_paciente', current_user.id)
                
                if new_nir.admission_date and new_nir.entry_type:
                    update_section_status(new_nir.id, 'dados_internacao_iniciais', current_user.id)
                
                if config_sections.get('agendamento_inicial') == 'NIR' and new_nir.scheduling_date:
                    update_section_status(new_nir.id, 'agendamento_inicial', current_user.id)
                
                if admission_type_form == 'CLINICO':
                    if config_sections.get('procedimentos') == 'NIR' and new_nir.procedure_code:
                        update_section_status(new_nir.id, 'procedimentos', current_user.id)
                    
                    if config_sections.get('informacoes_medicas') == 'NIR' and new_nir.responsible_doctor:
                        update_section_status(new_nir.id, 'informacoes_medicas', current_user.id)
                
                elif new_nir.entry_type == 'ELETIVO':
                    if config_sections.get('procedimentos') == 'NIR' and new_nir.procedure_code:
                        update_section_status(new_nir.id, 'procedimentos', current_user.id)
                    if config_sections.get('informacoes_medicas') == 'NIR' and new_nir.responsible_doctor:
                        update_section_status(new_nir.id, 'informacoes_medicas', current_user.id)
        except Exception as e:
            print(f"Erro ao marcar seções: {str(e)}")
            pass

        flash('Registro NIR criado com sucesso!', 'success')
        return redirect(url_for('nir.sector_nir_list'))

    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao criar registro: {str(e)}', 'danger')
        return redirect(url_for('nir.sector_new_record'))

#<!--- Detalhes do Registro NIR --->
@nir_bp.route("/nir/<int:record_id>")
@login_required
def record_details(record_id):
    record = Nir.query.get_or_404(record_id)
    header_wait_label = None
    header_wait_kind = None
    try:
        # Verifica se o registro está cancelado primeiro
        if (record.cancelled or '').upper() == 'SIM' or record.status == 'CANCELADO':
            header_wait_label = 'Fluxo cancelado'
            header_wait_kind = 'cancelled'
        # Verifica se está em observação
        elif record.status == 'EM_OBSERVACAO':
            header_wait_label = 'Em observação'
            header_wait_kind = 'observation'
        elif record.status == 'AGUARDANDO_DECISAO':
            header_wait_label = 'Aguardando decisão'
            header_wait_kind = 'observation'
        else:
            prog = record.get_sector_progress()
            billing_status = (prog.get('FATURAMENTO') or {}).get('status')
            if billing_status == 'CONCLUIDO':
                header_wait_label = 'Fluxo concluído'
                header_wait_kind = 'done'
            else:
                current_sector = record.get_next_available_sector()
                if current_sector == 'FATURAMENTO':
                    header_wait_label = 'Aguardando Faturamento'
                    header_wait_kind = 'billing'
                elif current_sector == 'CENTRO_CIRURGICO':
                    header_wait_label = 'Aguardando Centro Cirúrgico'
                    header_wait_kind = 'surgery'
                elif current_sector == 'NIR':
                    effective_entry = record.get_effective_entry_type()
                    config = record.get_section_control_config()
                    nir_sections = [s for s, sec in config.items() if sec == 'NIR']
                    alta_sections = [s for s in nir_sections if 'alta' in s]
                    initial_sections = [s for s in nir_sections if s not in alta_sections]
                    status_map = {s.section_name: s.status for s in record.section_statuses if s.responsible_sector == 'NIR'}
                    def sections_complete(lst):
                        if not lst:
                            return True
                        return all(status_map.get(sec) == 'PREENCHIDO' for sec in lst)
                    if effective_entry in ('URGENCIA', 'ELETIVO') and sections_complete(initial_sections) and not sections_complete(alta_sections):
                        header_wait_label = 'Aguardando dados de Alta'
                    else:
                        header_wait_label = 'Aguardando NIR'
    except Exception:
        pass
    return render_template('nir/record_details.html', record=record, header_wait_label=header_wait_label, header_wait_kind=header_wait_kind)

#<!--- Edição do Registro NIR --->
@nir_bp.route("/nir/editar/<int:record_id>")
@login_required
@require_permission('editar-registro-nir')
def edit_record(record_id):
    record = Nir.query.get_or_404(record_id)
    user_sector = get_user_sector(current_user)
    
    config = record.get_section_control_config()
    editable_sections = {}
    section_statuses = {}
    
    for section_name, responsible_sector in config.items():
        editable_sections[section_name] = True
        status_obj = record.get_section_status(section_name)
        section_statuses[section_name] = status_obj.status if status_obj else 'PENDENTE'
    
    return render_template('nir/edit_record.html', 
                         record=record, 
                         user_sector=user_sector,
                         editable_sections=editable_sections,
                         section_statuses=section_statuses,
                         config=config,
                         current_user=current_user)

#<!--- Atualização do Registro NIR --->
@nir_bp.route("/nir/atualizar/<int:record_id>", methods=['POST'])
@login_required
@require_permission('salvar-registro-nir')
def update_record(record_id):
    record = Nir.query.get_or_404(record_id)
    user_sector = get_user_sector(current_user)
    
    try:
        config = record.get_section_control_config()
        try:
            initialize_section_statuses(record)
        except Exception:
            pass
        editable_sections = {section_name: True for section_name in config.keys()}
        
        birth_date_str = request.form.get('birth_date')
        birth_date = datetime.strptime(birth_date_str, '%Y-%m-%d').date() if birth_date_str else None

        admission_date_str = request.form.get('admission_date')
        admission_date = parse_datetime_local(admission_date_str)

        scheduling_date_str = request.form.get('scheduling_date')
        scheduling_date = datetime.strptime(scheduling_date_str, '%Y-%m-%d').date() if scheduling_date_str else None

        discharge_date_str = request.form.get('discharge_date')
        discharge_date = parse_datetime_local(discharge_date_str)

        total_days = None
        if admission_date and discharge_date:
            total_days = (discharge_date.date() - admission_date.date()).days if hasattr(admission_date, 'date') else (discharge_date - admission_date).days

        changed_dados_paciente = False
        if editable_sections.get('dados_paciente', True):
            if 'patient_name' in request.form:
                record.patient_name = request.form.get('patient_name') or record.patient_name
                changed_dados_paciente = True
            if 'birth_date' in request.form:
                record.birth_date = birth_date or record.birth_date
                changed_dados_paciente = True
            if 'gender' in request.form:
                record.gender = request.form.get('gender') or record.gender
                changed_dados_paciente = True
            if 'susfacil' in request.form:
                record.susfacil = request.form.get('susfacil')
                changed_dados_paciente = True
            if 'sus_number' in request.form:
                if request.form.get('sus_number'):
                    record.sus_number = request.form.get('sus_number')
                changed_dados_paciente = True
            record.is_palliative = request.form.get('is_palliative') == 'on'
            changed_dados_paciente = True
        changed_dados_iniciais = False
        if editable_sections.get('dados_internacao_iniciais', True):
            if 'admission_date' in request.form:
                if admission_date:
                    record.admission_date = admission_date
                    record.day = admission_date.day
                    record.month = admission_date.strftime('%B')
                changed_dados_iniciais = True
            if 'entry_type' in request.form:
                if request.form.get('entry_type'):
                    record.entry_type = request.form.get('entry_type')
                changed_dados_iniciais = True
            if 'admission_type' in request.form:
                record.admission_type = request.form.get('admission_type')
                changed_dados_iniciais = True
            if 'admitted_from_origin' in request.form:
                record.admitted_from_origin = request.form.get('admitted_from_origin')
                changed_dados_iniciais = True
            if 'recurso' in request.form:
                record.recurso = request.form.get('recurso')
                changed_dados_iniciais = True
            if 'aih' in request.form:
                new_aih = request.form.get('aih')
                if new_aih and new_aih.strip():
                    existing_aih = Nir.query.filter_by(aih=new_aih.strip()).filter(Nir.id != record_id).first()
                    if existing_aih:
                        flash(f'Já existe uma solicitação com o AIH {new_aih}. O número de AIH deve ser único.', 'danger')
                        return redirect(url_for('nir.edit_record', record_id=record_id))
                    record.aih = new_aih
                changed_dados_iniciais = True

        changed_procedimentos = False
        if editable_sections.get('procedimentos', True):
            if any(k in request.form for k in ['procedure_codes[]','procedure_code','surgical_description']):
                record.procedures.clear()
                db.session.flush()
                codes = request.form.getlist('procedure_codes[]')
                descs = request.form.getlist('procedure_descriptions[]')
                seen = set()
                for idx, (c, d) in enumerate(zip(codes, descs)):
                    c_norm = (c or '').strip()
                    d_norm = (d or '').strip()
                    if not c_norm or not d_norm:
                        continue
                    if c_norm in seen:
                        continue
                    seen.add(c_norm)
                    record.procedures.append(NirProcedure(
                        code=c_norm,
                        description=d_norm,
                        sequence=idx,
                        is_primary=(idx == 0)
                    ))
                    if idx == 0:
                        record.procedure_code = c_norm
                        record.surgical_description = d_norm

                if not seen:
                    single_code = request.form.get('procedure_code')
                    single_desc = request.form.get('surgical_description')
                    if single_code:
                        record.procedure_code = single_code
                    if single_desc:
                        record.surgical_description = single_desc
                    if single_code and single_desc:
                        record.procedures.append(NirProcedure(
                            code=single_code,
                            description=single_desc,
                            sequence=0,
                            is_primary=True
                        ))
                changed_procedimentos = True

        changed_info_med = False
        if editable_sections.get('informacoes_medicas', True):
            if 'responsible_doctor' in request.form:
                record.responsible_doctor = request.form.get('responsible_doctor') or record.responsible_doctor
                changed_info_med = True
            if 'main_cid' in request.form:
                record.main_cid = request.form.get('main_cid') or record.main_cid
                changed_info_med = True
            if 'surgical_specialty' in request.form:
                record.surgical_specialty = request.form.get('surgical_specialty') or record.surgical_specialty
                changed_info_med = True
            if 'auxiliary' in request.form:
                record.auxiliary = request.form.get('auxiliary') or record.auxiliary
                changed_info_med = True
            if 'anesthetist' in request.form:
                record.anesthetist = request.form.get('anesthetist') or record.anesthetist
                changed_info_med = True
            if 'anesthesia' in request.form:
                record.anesthesia = request.form.get('anesthesia') or record.anesthesia
                changed_info_med = True
            if 'pediatrics' in request.form:
                if request.form.get('pediatrics_nao_aplica'):
                    record.pediatrics = None
                else:
                    record.pediatrics = request.form.get('pediatrics') or record.pediatrics
                changed_info_med = True
            if 'surgical_type' in request.form:
                record.surgical_type = request.form.get('surgical_type') or record.surgical_type
                changed_info_med = True
            if 'auxiliary_nao_aplica' in request.form and request.form.get('auxiliary_nao_aplica'):
                record.auxiliary = None
                changed_info_med = True

        changed_agendamento = False
        if editable_sections.get('agendamento_inicial', True):
            if 'scheduling_date' in request.form:
                if scheduling_date:
                    record.scheduling_date = scheduling_date
                changed_agendamento = True

        changed_alta = False
        if editable_sections.get('dados_alta_finais', True):
            if 'discharge_type' in request.form:
                record.discharge_type = request.form.get('discharge_type')
                changed_alta = True
            if 'discharge_date' in request.form:
                record.discharge_date = discharge_date
                if discharge_date:
                    admission_for_calc = admission_date or record.admission_date
                    if admission_for_calc:
                        record.total_days_admitted = (discharge_date - admission_for_calc).days
                changed_alta = True
            if 'aih_final' in request.form:
                aih_final = request.form.get('aih_final')
                if aih_final:
                    record.aih = aih_final
                changed_alta = True

        changed_status = False
        if editable_sections.get('status_controle', True):
            if 'status' in request.form:
                status_final = (request.form.get('status') or '').strip()
                if status_final == 'CANCELADO':
                    record.cancelled = 'SIM'
                elif status_final:
                    record.cancelled = 'NAO'
                changed_status = True
            if 'cancellation_reason' in request.form:
                record.cancellation_reason = request.form.get('cancellation_reason')
                changed_status = True
            def norm_bool(value):
                if not value:
                    return None
                value_up = value.upper()
                if value_up in ('SIM', 'NAO', 'NÃO'):
                    return 'SIM' if value_up == 'SIM' else 'NAO'
                return None
            if 'criticized' in request.form:
                record.criticized = norm_bool((request.form.get('criticized') or '').strip()) or record.criticized
                changed_status = True
            if 'billed' in request.form:
                record.billed = norm_bool((request.form.get('billed') or '').strip()) or record.billed
                changed_status = True
            if 'observation' in request.form:
                record.observation = request.form.get('observation')
                changed_status = True

        record.last_modified = datetime.utcnow()

        try:
            if record.admission_date and record.discharge_date:
                record.total_days_admitted = (record.discharge_date - record.admission_date).days
        except Exception:
            pass
        
        if editable_sections.get('dados_paciente', False) and changed_dados_paciente:
            update_section_status(record.id, 'dados_paciente', current_user.id)
        if editable_sections.get('dados_internacao_iniciais', False) and changed_dados_iniciais:
            update_section_status(record.id, 'dados_internacao_iniciais', current_user.id)
        if editable_sections.get('agendamento_inicial', False) and changed_agendamento:
            update_section_status(record.id, 'agendamento_inicial', current_user.id)
        if editable_sections.get('procedimentos', False) and changed_procedimentos:
            update_section_status(record.id, 'procedimentos', current_user.id)
        if editable_sections.get('informacoes_medicas', False) and changed_info_med:
            update_section_status(record.id, 'informacoes_medicas', current_user.id)
        if editable_sections.get('dados_alta_finais', False) and changed_alta:
            update_section_status(record.id, 'dados_alta_finais', current_user.id)
        if editable_sections.get('status_controle', False) and changed_status:
            update_section_status(record.id, 'status_controle', current_user.id)

        try:
            record.status = record.compute_overall_status()
        except Exception:
            pass

        db.session.commit()
        
        allowed_sections = [name for name, can_edit in editable_sections.items() if can_edit]
        flash(f'Seções atualizadas: {", ".join(allowed_sections)}', 'success')
        return redirect(url_for('nir.record_details', record_id=record.id))
        
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao atualizar registro: {str(e)}', 'danger')
        return redirect(url_for('nir.edit_record', record_id=record_id))

#<!--- Exclusão do Registro NIR --->
@nir_bp.route("/nir/excluir/<int:record_id>", methods=['POST'])
@login_required
@require_permission('excluir-registro-nir')
def delete_record(record_id):
    record = Nir.query.get_or_404(record_id)

    try:
        db.session.delete(record)
        db.session.commit()
        flash('Registro NIR excluído com sucesso!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao excluir registro: {str(e)}', 'danger')

    return redirect(url_for('nir.list_records'))

#<!--- Rota de Redirecionamento "Meus Trabalhos" --->
@nir_bp.route("/nir/meus-trabalhos")
@login_required 
def my_work():
    user_sector = get_user_sector(current_user)
    
    if user_sector == 'NIR':
        return redirect(url_for('nir.sector_nir_list'))
    elif user_sector == 'CENTRO_CIRURGICO':
        return redirect(url_for('nir.sector_surgery_list'))
    elif user_sector == 'FATURAMENTO':
        return redirect(url_for('nir.sector_billing_list'))
    else:
        return redirect(url_for('nir.sector_nir_list'))

#<!--- Rota de Listagem do Setor NIR --->
@nir_bp.route("/nir/setor/nir")
@login_required 
@require_sector('NIR')
def sector_nir_list():
    user_sector = get_user_sector(current_user)
    if user_sector != 'NIR':
        flash('Acesso negado: você não pertence ao setor NIR', 'danger')
        return redirect(url_for('nir.my_work'))
    
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 10, type=int)
    if per_page <= 0:
        per_page = 10
    if per_page > 100:
        per_page = 100

    filter_status = request.args.get('filter_status', '').strip().lower()
    valid_filters = {'pendente': 'PENDENTE', 'andamento': 'EM_ANDAMENTO', 'concluido': 'CONCLUIDO', 'observacao': 'OBSERVACAO', 'aguardando_decisao': 'AGUARDANDO_DECISAO'}
    target_status = valid_filters.get(filter_status)
    waiting_for_param = request.args.get('waiting_for', '').strip().lower()
    valid_waiting = {'faturamento': 'FATURAMENTO', 'cirurgia': 'CENTRO CIRÚRGICO'}
    target_waiting = valid_waiting.get(waiting_for_param)
    
    query = Nir.query.order_by(Nir.creation_date.desc())
    
    patient_name = request.args.get('patient_name', '').strip()
    if patient_name:
        query = query.filter(Nir.patient_name.ilike(f'%{patient_name}%'))
    
    all_records = query.all()
    nir_relevant_records = []
    
    for record in all_records:
        if record.status == 'CANCELADO':
            continue
        
        if record.status in ('EM_OBSERVACAO', 'AGUARDANDO_DECISAO'):
            if record.status == 'EM_OBSERVACAO' and record.should_transition_to_decision():
                try:
                    record.status = 'AGUARDANDO_DECISAO'
                    db.session.commit()
                except Exception as e:
                    db.session.rollback()
                    print(f"Erro ao atualizar status de observação: {e}")
            
            setattr(record, '_created_by_current', record.operator_id == current_user.id)
            setattr(record, '_nir_phase', 'OBSERVACAO')
            setattr(record, '_nir_locked', False)
            
            if record.status == 'AGUARDANDO_DECISAO':
                setattr(record, '_nir_display_status', 'PENDENTE')
            else:
                setattr(record, '_nir_display_status', 'EM_OBSERVACAO')
            
            setattr(record, '_nir_waiting_for', None)
            setattr(record, '_is_observation', True)
            nir_relevant_records.append(record)
            continue
        
        progress = record.get_sector_progress()
        nir_sector_progress = progress.get('NIR', {})
        entry = record.get_effective_entry_type()
        config = record.get_section_control_config()
        nir_sections = [s for s, sec in config.items() if sec == 'NIR']
        alta_sections = [s for s in nir_sections if 'alta' in s]
        initial_sections = [s for s in nir_sections if s not in alta_sections]
        status_map = {s.section_name: s.status for s in record.section_statuses if s.responsible_sector == 'NIR'}
        def sections_complete(lst):
            if not lst:
                return True
            return all(status_map.get(sec) == 'PREENCHIDO' for sec in lst)
        initial_complete = sections_complete(initial_sections)
        alta_complete = sections_complete(alta_sections)
        surgery_progress = progress.get('CENTRO_CIRURGICO')
        surgery_complete = (surgery_progress and surgery_progress.get('status') == 'CONCLUIDO') if surgery_progress else True

        setattr(record, '_created_by_current', record.operator_id == current_user.id)
        setattr(record, '_is_observation', False)
        try:
            progress = record.get_sector_progress()
            billing_progress = progress.get('FATURAMENTO', {})
            billing_complete = billing_progress.get('status') == 'CONCLUIDO'
            
            phase_info = get_nir_phase(record)
            phase = phase_info.get('phase')
            is_locked = phase_info.get('locked', False)
            setattr(record, '_nir_phase', phase)
            setattr(record, '_nir_locked', is_locked)
            
            display_status = nir_sector_progress.get('status')
            
            if billing_complete:
                display_status = 'CONCLUIDO'
            elif is_locked:
                if nir_sector_progress.get('filled', 0) > 0:
                    if display_status == 'CONCLUIDO':
                        display_status = 'CONCLUIDO'
                    else:
                        display_status = 'EM_ANDAMENTO'
            else:
                if phase in ('INITIAL', 'FULL'):
                    display_status = 'PENDENTE'
                elif phase == 'FINAL':
                    if not alta_complete:
                        if not record.discharge_date and not record.discharge_type:
                            display_status = 'PENDENTE'
            
            setattr(record, '_nir_display_status', display_status)
            setattr(record, '_nir_waiting_for', phase_info.get('waiting_for'))
        except Exception:
            setattr(record, '_nir_phase', None)
            setattr(record, '_nir_display_status', nir_sector_progress.get('status'))
            setattr(record, '_nir_waiting_for', None)
            setattr(record, '_nir_locked', False)
        nir_relevant_records.append(record)
    
    if waiting_for_param == 'alta':
        nir_relevant_records = [r for r in nir_relevant_records if getattr(r, '_nir_phase', None) == 'FINAL']
    elif target_waiting:
        nir_relevant_records = [r for r in nir_relevant_records if getattr(r, '_nir_waiting_for', None) == target_waiting]

    records_for_counts = list(nir_relevant_records)

    def _status_priority(rec):
        display = getattr(rec, '_nir_display_status', None) or rec.get_sector_progress().get('NIR', {}).get('status')
        is_observation = getattr(rec, '_is_observation', False)
        
        if is_observation and rec.status == 'AGUARDANDO_DECISAO':
            pri = -1
        elif display == 'PENDENTE':
            pri = 0
        elif display == 'EM_ANDAMENTO':
            pri = 1
        elif is_observation or display in ('EM_OBSERVACAO', 'AGUARDANDO_DECISAO'):
            pri = 1.5
        else:
            pri = 2
        
        created_order = rec.creation_date or datetime.min
        return (pri, -created_order.timestamp())

    try:
        nir_relevant_records.sort(key=_status_priority)
    except Exception:
        nir_relevant_records.sort(key=lambda r: 0)

    stats_counts = {
        'total': len(nir_relevant_records),
        'pendentes': 0,
        'andamento': 0,
        'concluidos': 0,
        'meus': 0,
        'wait_billing': 0,
        'wait_surgery': 0,
        'observacoes': 0,
        'aguardando_decisao': 0,
    }
    for r in records_for_counts:
        if getattr(r, '_created_by_current', False):
            stats_counts['meus'] += 1
        display = getattr(r, '_nir_display_status', None) or r.get_sector_progress().get('NIR', {}).get('status')
        
        if display in ('EM_OBSERVACAO', 'AGUARDANDO_DECISAO'):
            if display == 'EM_OBSERVACAO':
                stats_counts['observacoes'] += 1
            else:
                stats_counts['aguardando_decisao'] += 1
        elif display == 'PENDENTE':
            stats_counts['pendentes'] += 1
        elif display == 'EM_ANDAMENTO':
            stats_counts['andamento'] += 1
        elif display == 'CONCLUIDO':
            stats_counts['concluidos'] += 1
            
        waiting_val = getattr(r, '_nir_waiting_for', None)
        if waiting_val == 'FATURAMENTO':
            stats_counts['wait_billing'] += 1
        elif waiting_val == 'CENTRO CIRÚRGICO':
            stats_counts['wait_surgery'] += 1

    if target_status:
        if target_status == 'OBSERVACAO':
            # Filtra apenas observações < 24h (status EM_OBSERVACAO)
            nir_relevant_records = [
                r for r in nir_relevant_records
                if getattr(r, '_is_observation', False) and r.status == 'EM_OBSERVACAO'
            ]
        elif target_status == 'AGUARDANDO_DECISAO':
            # Filtra apenas observações > 24h (status AGUARDANDO_DECISAO)
            nir_relevant_records = [
                r for r in nir_relevant_records
                if getattr(r, '_is_observation', False) and r.status == 'AGUARDANDO_DECISAO'
            ]
        else:
            nir_relevant_records = [
                r for r in nir_relevant_records
                if (getattr(r, '_nir_display_status', None) or r.get_sector_progress().get('NIR', {}).get('status')) == target_status
            ]

    total = len(nir_relevant_records)
    start_idx = (page - 1) * per_page
    end_idx = start_idx + per_page
    page_items = nir_relevant_records[start_idx:end_idx]
    pages = (total // per_page + (1 if total % per_page else 0)) if total > 0 else 1

    from types import SimpleNamespace
    pagination = SimpleNamespace(
        items=page_items,
        total=total,
        pages=pages,
        page=page,
        has_prev=page > 1,
        has_next=page < pages,
        prev_num=page - 1 if page > 1 else None,
        next_num=page + 1 if page < pages else None,
        iter_pages=create_iter_pages_function(pages, page)
    )

    records = pagination
    
    return render_template(
        'nir/sector_nir_list.html', 
        records=records, 
        user_sector=user_sector, 
        filter_status=filter_status, 
        waiting_for=waiting_for_param,
    pagination=pagination,
    stats_counts=stats_counts
    )

#<!--- Rota de Listagem do Setor Centro Cirúrgico --->
@nir_bp.route("/nir/setor/centro-cirurgico")
@login_required
@require_sector('CENTRO_CIRURGICO')
def sector_surgery_list():
    user_sector = get_user_sector(current_user)
    if user_sector != 'CENTRO_CIRURGICO':
        flash('Acesso negado: você não pertence ao Centro Cirúrgico', 'danger')
        return redirect(url_for('nir.my_work'))
    
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 10, type=int)
    if per_page <= 0:
        per_page = 10
    if per_page > 100:
        per_page = 100

    query = Nir.query.order_by(Nir.creation_date.desc())
    
    patient_name = request.args.get('patient_name', '').strip()
    if patient_name:
        query = query.filter(Nir.patient_name.ilike(f'%{patient_name}%'))
        
    all_records = query.all()
    surgery_relevant_records = []
    
    for record in all_records:
        progress = record.get_sector_progress()
        surgery_progress = progress.get('CENTRO_CIRURGICO', {})
        status = surgery_progress.get('status')

        include = False
        if record.is_ready_for_sector('CENTRO_CIRURGICO'):
            include = status in ['PENDENTE', 'EM_ANDAMENTO']
        if status == 'CONCLUIDO':
            include = True

        if include:
            surgery_relevant_records.append(record)

    filter_status = request.args.get('filter_status', '').strip().lower()
    status_map = {'pendente': 'PENDENTE', 'andamento': 'EM_ANDAMENTO', 'concluido': 'CONCLUIDO'}
    target_status = status_map.get(filter_status)

    def _status_priority(rec):
        prog = rec.get_sector_progress().get('CENTRO_CIRURGICO', {})
        display = prog.get('status')
        pri = 2
        if display == 'PENDENTE':
            pri = 0
        elif display == 'EM_ANDAMENTO':
            pri = 1
        created_order = rec.creation_date or datetime.min
        return (pri, -created_order.timestamp())
    try:
        surgery_relevant_records.sort(key=_status_priority)
    except Exception:
        surgery_relevant_records.sort(key=lambda r: 0)

    stats_counts = {
        'total': len(surgery_relevant_records),
        'pendentes': 0,
        'andamento': 0,
        'concluidos': 0,
    }
    for r in surgery_relevant_records:
        st = r.get_sector_progress().get('CENTRO_CIRURGICO', {}).get('status')
        if st == 'PENDENTE':
            stats_counts['pendentes'] += 1
        elif st == 'EM_ANDAMENTO':
            stats_counts['andamento'] += 1
        elif st == 'CONCLUIDO':
            stats_counts['concluidos'] += 1

    if target_status:
        surgery_relevant_records = [
            r for r in surgery_relevant_records
            if r.get_sector_progress().get('CENTRO_CIRURGICO', {}).get('status') == target_status
        ]

    total = len(surgery_relevant_records)
    start_idx = (page - 1) * per_page
    end_idx = start_idx + per_page
    page_items = surgery_relevant_records[start_idx:end_idx]
    pages = (total // per_page + (1 if total % per_page else 0)) if total > 0 else 1

    from types import SimpleNamespace
    pagination = SimpleNamespace(
        items=page_items,
        total=total,
        pages=pages,
        page=page,
        has_prev=page > 1,
        has_next=page < pages,
        prev_num=page - 1 if page > 1 else None,
        next_num=page + 1 if page < pages else None,
        iter_pages=create_iter_pages_function(pages, page)
    )

    return render_template(
        'nir/sector_surgery_list.html', 
        records=pagination, 
        user_sector=user_sector,
        pagination=pagination,
    stats_counts=stats_counts,
    filter_status=filter_status
    )

#<!--- Rota de Listagem do Setor Faturamento --->
@nir_bp.route("/nir/setor/faturamento")
@login_required
@require_sector('FATURAMENTO')
def sector_billing_list():
    user_sector = get_user_sector(current_user)
    if user_sector != 'FATURAMENTO':
        flash('Acesso negado: você não pertence ao Faturamento', 'danger')
        return redirect(url_for('nir.my_work'))
    
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 10, type=int)
    if per_page <= 0:
        per_page = 10
    if per_page > 100:
        per_page = 100

    query = Nir.query.order_by(Nir.creation_date.desc())
    
    patient_name = request.args.get('patient_name', '').strip()
    if patient_name:
        query = query.filter(Nir.patient_name.ilike(f'%{patient_name}%'))

    all_records = query.all()
    billing_relevant_records = []
    
    for record in all_records:
        progress = record.get_sector_progress()
        billing_progress = progress.get('FATURAMENTO', {})
        status = billing_progress.get('status')

        include = False
        if record.is_ready_for_sector('FATURAMENTO'):
            include = status in ['PENDENTE', 'EM_ANDAMENTO']
        if status == 'CONCLUIDO':
            include = True

        if include:
            billing_relevant_records.append(record)

    filter_status = request.args.get('filter_status', '').strip().lower()
    status_map = {'pendente': 'PENDENTE', 'andamento': 'EM_ANDAMENTO', 'concluido': 'CONCLUIDO'}
    target_status = status_map.get(filter_status)

    def _status_priority(rec):
        st = rec.get_sector_progress().get('FATURAMENTO', {}).get('status')
        pri = 2
        if st == 'PENDENTE':
            pri = 0
        elif st == 'EM_ANDAMENTO':
            pri = 1
        created_order = rec.creation_date or datetime.min
        return (pri, -created_order.timestamp())
    try:
        billing_relevant_records.sort(key=_status_priority)
    except Exception:
        billing_relevant_records.sort(key=lambda r: 0)

    stats_counts = {
        'total': len(billing_relevant_records),
        'pendentes': 0,
        'andamento': 0,
        'concluidos': 0,
    }
    for r in billing_relevant_records:
        st = (r.get_sector_progress().get('FATURAMENTO', {}) or {}).get('status')
        if st == 'PENDENTE':
            stats_counts['pendentes'] += 1
        elif st == 'EM_ANDAMENTO':
            stats_counts['andamento'] += 1
        elif st == 'CONCLUIDO':
            stats_counts['concluidos'] += 1

    if target_status:
        filtered_records = [
            r for r in billing_relevant_records
            if (r.get_sector_progress().get('FATURAMENTO', {}) or {}).get('status') == target_status
        ]
    else:
        filtered_records = billing_relevant_records

    total = len(filtered_records)
    start_idx = (page - 1) * per_page
    end_idx = start_idx + per_page
    page_items = filtered_records[start_idx:end_idx]
    pages = (total // per_page + (1 if total % per_page else 0)) if total > 0 else 1

    from types import SimpleNamespace
    pagination = SimpleNamespace(
        items=page_items,
        total=total,
        pages=pages,
        page=page,
        has_prev=page > 1,
        has_next=page < pages,
        prev_num=page - 1 if page > 1 else None,
        next_num=page + 1 if page < pages else None,
        iter_pages=create_iter_pages_function(pages, page)
    )
    
    return render_template(
        'nir/sector_billing_list.html',
        records=pagination,
        user_sector=user_sector,
        pagination=pagination,
        stats_counts=stats_counts,
        filter_status=filter_status
    )

#<!--- Rota de Preencher Formulário do Setor NIR --->
@nir_bp.route("/nir/<int:record_id>/setor/nir")
@login_required
@require_sector('NIR')
def sector_nir_form(record_id):
    record = Nir.query.get_or_404(record_id)
    user_sector = get_user_sector(current_user)
    
    section_status_map = {s.section_name: s.status for s in record.section_statuses if s.responsible_sector == 'NIR'}
    phase_info = get_nir_phase(record)
    final_phase = phase_info['phase'] == 'FINAL'
    
    hide_aih_initial = phase_info['phase'] in ('INITIAL', 'LOCKED_WAIT_SURGERY')

    config = record.get_section_control_config()
    nir_sections = [s for s, sec in config.items() if sec == 'NIR']
    
    initial_sections = ['dados_paciente', 'dados_internacao_iniciais', 'agendamento_inicial']
    alta_sections = ['dados_alta_finais']
    
    initial_sections = [s for s in initial_sections if s in nir_sections]
    alta_sections = [s for s in alta_sections if s in nir_sections]

    display_sections = []
    editable_sections = set(phase_info['show_sections']) if phase_info['show_sections'] else set()
    if phase_info['locked']:
        if phase_info['phase'] == 'LOCKED_WAIT_SURGERY':
            display_sections = initial_sections + alta_sections
        elif phase_info['phase'] in ('LOCKED_AFTER','LOCKED_AFTER_FULL'):
            display_sections = nir_sections
        else:
            display_sections = nir_sections
    else:
        display_sections = phase_info['show_sections']

    return render_template(
        'nir/forms/nir_sector_form.html',
        record=record,
        user_sector=user_sector,
        final_phase=final_phase,
        show_sections=display_sections,
        editable_sections=list(editable_sections),
        section_status_map=section_status_map,
        phase_info=phase_info,
        config=config,
        hide_aih_initial=hide_aih_initial
    )

#<!--- Rota de Preencher Formulário do Setor Faturamento --->
@nir_bp.route("/nir/<int:record_id>/setor/faturamento")
@login_required
@require_sector('FATURAMENTO')
def sector_billing_form(record_id):
    record = Nir.query.get_or_404(record_id)
    user_sector = get_user_sector(current_user)
    
    if not record.is_ready_for_sector('FATURAMENTO'):
        next_sector = record.get_next_available_sector()
        if next_sector == 'NIR':
            flash('Finalize também os dados de alta no NIR antes do faturamento.', 'warning')
        elif next_sector == 'CENTRO_CIRURGICO':
            flash('Aguarde a conclusão do Centro Cirúrgico antes do faturamento.', 'warning')
        else:
            flash('Este registro não está pronto para o Faturamento.', 'warning')
        return redirect(url_for('nir.record_details', record_id=record_id))
    
    try:
        progress = record.get_sector_progress()
        billing_status = (progress.get('FATURAMENTO') or {}).get('status')
        if billing_status == 'CONCLUIDO':
            flash('Este registro do Faturamento já está concluído e não pode ser editado.', 'info')
            return redirect(url_for('nir.record_details', record_id=record_id))
    except Exception:
        pass
    
    return render_template('nir/forms/billing_sector_form.html', record=record, user_sector=user_sector)

#<!--- Rota de Busca de Procedimentos --->
@nir_bp.route('/nir/procedures/search')
@login_required
def procedures_search():
    def remove_accents(text):
        if not text:
            return ''
        nfkd = unicodedata.normalize('NFD', text)
        return ''.join([c for c in nfkd if not unicodedata.category(c).startswith('M')])
    
    q = request.args.get('q', '').strip()
    limit = request.args.get('limit', 15, type=int)
    if limit > 50:
        limit = 50
    
    if not q:
        return jsonify([])
    
    procedures = Procedure.query.filter(Procedure.is_active == True).all()
    
    q_normalized = remove_accents(q.lower())
    
    filtered_procedures = []
    for p in procedures:
        code_normalized = remove_accents(p.code.lower())
        desc_normalized = remove_accents(p.description.lower())
        
        if q_normalized in code_normalized or q_normalized in desc_normalized:
            filtered_procedures.append(p)
            if len(filtered_procedures) >= limit:
                break
    
    filtered_procedures.sort(key=lambda x: x.code)
    
    results = [{'code': p.code, 'description': p.description} for p in filtered_procedures]
    return jsonify(results)

#<!--- Rota de Busca de CIDs Relacionados a um Procedimento --->
@nir_bp.route('/nir/procedure/<procedure_code>/cids')
@login_required
def get_procedure_cids(procedure_code):
    try:
        from app.procedures_models import ProcedureCid, Cid
        
        relationships = ProcedureCid.query.filter_by(procedure_code=procedure_code).all()
        cid_codes = [rel.cid_code for rel in relationships]
        
        if not cid_codes:
            return jsonify([])
        
        cids = Cid.query.filter(
            Cid.code.in_(cid_codes),
            Cid.is_active == True
        ).all()
        
        results = [{'code': c.code, 'description': c.description} for c in cids]
        return jsonify(results)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

#<!--- Exportar Registros NIR para Excel --->
@nir_bp.route("/nir/exportar-excel")
@login_required
def export_to_excel():
    try:
        search = request.args.get('search', '').strip()
        entry_type = request.args.get('entry_type', '').strip()
        admission_type = request.args.get('admission_type', '').strip()
        is_palliative = request.args.get('is_palliative', '').strip()
        origin = request.args.get('origin', '').strip()
        recurso = request.args.get('recurso', '').strip()
        responsible_doctor = request.args.get('responsible_doctor', '').strip()
        start_date = request.args.get('start_date', '').strip()
        end_date = request.args.get('end_date', '').strip()
        sector_progress = request.args.get('sector_progress', '').strip()

        query = Nir.query.order_by(Nir.creation_date.desc())

        if search:
            query = query.filter(
                (Nir.patient_name.ilike(f'%{search}%')) |
                (Nir.susfacil_protocol.ilike(f'%{search}%'))
            )

        if entry_type:
            query = query.filter(Nir.entry_type.ilike(entry_type))

        if admission_type:
            query = query.filter(Nir.admission_type.ilike(admission_type))

        if discharge_type:
            query = query.filter(Nir.discharge_type.ilike(discharge_type))
        
        if is_palliative:
            if is_palliative == '1':
                query = query.filter(Nir.is_palliative == True)
            elif is_palliative == '0':
                query = query.filter(Nir.is_palliative == False)
        
        if origin:
            query = query.filter(Nir.admitted_from_origin == origin)
        
        if recurso:
            query = query.filter(Nir.recurso == recurso)

        if responsible_doctor:
            query = query.filter(Nir.responsible_doctor.ilike(f'%{responsible_doctor}%'))

        if start_date:
            try:
                start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
                query = query.filter(Nir.admission_date >= start_date_obj)
            except ValueError:
                pass

        if end_date:
            try:
                end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
                query = query.filter(Nir.admission_date <= end_date_obj)
            except ValueError:
                pass

        all_records = query.all()

        for rec in all_records:
            try:
                st, hint = _compute_global_info(rec)
                setattr(rec, '_global_status', st)
                setattr(rec, '_global_status_hint', hint)
            except Exception:
                setattr(rec, '_global_status', None)
                setattr(rec, '_global_status_hint', None)

        if sector_progress:
            display_records = [r for r in all_records if (getattr(r, '_global_status', None) or 'PENDENTE') == sector_progress]
        else:
            display_records = all_records

        wb = Workbook()
        ws = wb.active
        ws.title = "Registros NIR"

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        border_thin = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        headers = [
            'ID', 'Nome do Paciente', 'Data Nascimento', 'Gênero',
            'SUSFacil', 'Número SUS', 'Paliativo', 'Protocolo SUSFACIL', 
            'Data Internação', 'Tipo de Entrada', 'Tipo de Internação',
            'Origem', 'Recurso', 'Código Procedimento', 'Descrição Cirúrgica', 'Médico Responsável',
            'CID Principal', 'AIH', 'Data Agendamento', 'Tipo Alta', 'Data Alta',
            'Dias Internado', 'Especialidade Cirúrgica', 'Auxiliar', 'Anestesista',
            'Anestesia', 'Pediatria', 'Tipo Cirúrgico', 'Status', 'Cancelado',
            'Motivo Cancelamento', 'Criticado', 'Faturado', 'Observação', 
            'Status do Fluxo', 'Situação Atual', 'Data Criação'
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border_thin

        for row_num, record in enumerate(display_records, 2):
            procedures_text = ""
            if record.procedures:
                procedures_text = "; ".join([f"{p.code} - {p.description}" for p in record.procedures])
            elif record.procedure_code:
                procedures_text = f"{record.procedure_code} - {record.surgical_description or ''}"

            data = [
                record.id,
                record.patient_name or '',
                record.birth_date.strftime('%d/%m/%Y') if record.birth_date else '',
                record.gender or '',
                record.susfacil or '',
                record.sus_number or '',
                'Sim' if record.is_palliative else 'Não',
                record.susfacil_protocol or '',
                format_date_filter(record.admission_date, format_str='%d/%m/%Y %H:%M') if record.admission_date else '',
                record.entry_type or '',
                record.admission_type or '',
                record.admitted_from_origin or '',
                record.recurso or '',
                record.procedure_code or '',
                procedures_text,
                record.responsible_doctor or '',
                record.main_cid or '',
                record.aih or '',
                format_date_filter(record.scheduling_date, format_str='%d/%m/%Y') if record.scheduling_date else '',
                record.discharge_type or '',
                format_date_filter(record.discharge_date, format_str='%d/%m/%Y %H:%M') if record.discharge_date else '',
                record.total_days_admitted or '',
                record.surgical_specialty or '',
                record.auxiliary or '',
                record.anesthetist or '',
                record.anesthesia or '',
                record.pediatrics or '',
                record.surgical_type or '',
                record.status or '',
                record.cancelled or '',
                record.cancellation_reason or '',
                record.criticized or '',
                record.billed or '',
                record.observation or '',
                getattr(record, '_global_status', '') or '',
                getattr(record, '_global_status_hint', '') or '',
                format_date_filter(record.creation_date, format_str='%d/%m/%Y %H:%M') if record.creation_date else ''
            ]

            for col_num, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = border_thin
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        column_widths = {
            1: 8,   # ID
            2: 30,  # Nome Paciente
            3: 15,  # Data Nasc
            4: 10,  # Gênero
            5: 18,  # SUSFacil
            6: 18,  # SUS
            7: 25,  # Protocolo
            8: 15,  # Data Internação
            9: 18,  # Tipo Entrada
            10: 20, # Tipo Internação
            11: 15, # Origem
            12: 18, # Código Proc
            13: 50, # Descrição
            14: 25, # Médico
            15: 12, # CID
            16: 15, # AIH
            17: 15, # Data Agend
            18: 15, # Tipo Alta
            19: 15, # Data Alta
            20: 12, # Dias
            21: 20, # Especialidade
            22: 20, # Auxiliar
            23: 20, # Anestesista
            24: 15, # Anestesia
            25: 15, # Pediatria
            26: 15, # Tipo Cirúrgico
            27: 15, # Status
            28: 12, # Cancelado
            29: 30, # Motivo Cancel
            30: 12, # Criticado
            31: 12, # Faturado
            32: 40, # Observação
            33: 18, # Status Fluxo
            34: 25, # Situação
            35: 18  # Data Criação
        }

        for col_num, width in column_widths.items():
            ws.column_dimensions[get_column_letter(col_num)].width = width

        ws.freeze_panes = 'A2'

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'relatorio_nir_{timestamp}.xlsx'

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        flash(f'Erro ao exportar relatório: {str(e)}', 'danger')
        return redirect(url_for('nir.list_records'))

#<!--- Rota de Importação do SIGTAP --->
@nir_bp.route("/nir/import-sigtap", methods=['POST'])
@login_required
@require_permission('manage_procedures')
def import_sigtap():
    from app.utils.sigtap_importer import SIGTAPImporter
    import tempfile
    
    try:
        if 'sigtap_file' not in request.files:
            flash('Nenhum arquivo foi selecionado.', 'danger')
            return redirect(url_for('nir.sector_billing_list'))
        
        file = request.files['sigtap_file']
        
        if file.filename == '':
            flash('Nenhum arquivo foi selecionado.', 'danger')
            return redirect(url_for('nir.sector_billing_list'))
        
        if not file.filename.lower().endswith('.zip'):
            flash('Por favor, envie um arquivo .zip do SIGTAP.', 'danger')
            return redirect(url_for('nir.sector_billing_list'))
        
        filename = secure_filename(file.filename)
        temp_dir = tempfile.gettempdir()
        temp_path = os.path.join(temp_dir, filename)
        file.save(temp_path)
        
        try:
            importer = SIGTAPImporter()
            stats = importer.import_from_zip(temp_path)
            
            success_msg = f"""
            Importação do SIGTAP concluída com sucesso!<br>
            <strong>Procedimentos:</strong><br>
            • Novos: {stats['procedures']['inserted']}<br>
            • Atualizados: {stats['procedures']['updated']}<br>
            • Desativados: {stats['procedures']['deactivated']}<br>
            <strong>CIDs:</strong><br>
            • Novos: {stats['cids']['inserted']}<br>
            • Atualizados: {stats['cids']['updated']}<br>
            • Desativados: {stats['cids']['deactivated']}<br>
            <strong>Relacionamentos:</strong><br>
            • Inseridos: {stats['relationships']['inserted']}<br>
            """
            
            total_errors = stats['procedures']['errors'] + stats['cids']['errors'] + stats['relationships']['errors']
            if total_errors > 0:
                success_msg += f"<br>•  Erros encontrados: {total_errors}"
                
                if stats['error_messages']:
                    success_msg += "<br><br><strong>Primeiros erros:</strong><br>"
                    for error in stats['error_messages'][:5]:
                        success_msg += f"• {error}<br>"
                    
                    if len(stats['error_messages']) > 5:
                        success_msg += f"<br>... e mais {len(stats['error_messages']) - 5} erros."
                
                flash(success_msg, 'warning')
            else:
                flash(success_msg, 'success')
        
        finally:
            if os.path.exists(temp_path):
                os.remove(temp_path)
        
        return redirect(url_for('nir.sector_billing_list'))
    
    except FileNotFoundError as e:
        flash(f'Arquivo do SIGTAP não encontrado no ZIP: {str(e)}', 'danger')
        return redirect(url_for('nir.sector_billing_list'))
    
    except Exception as e:
        flash(f'Erro ao importar SIGTAP: {str(e)}', 'danger')
        return redirect(url_for('nir.sector_billing_list'))